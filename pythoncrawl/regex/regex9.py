import re
import openpyxl

# train.xlsx 데이터를 읽어 들여 4개의 시트를 만들어
# 각 행의 정보를 복사하기
# Mr (시트명은 남성)
# Miss (시트명은 미혼여성)
# Mrs (시트명 : 기혼여성)
# Others (시트명 : 기타)

excel_file = openpyxl.load_workbook("./resources/train.xlsx")
sheet1 = excel_file.active

wb = openpyxl.Workbook()
sheet1_man = wb.active
sheet1_man.column_dimensions["D"].width = 70
sheet1_man.title = "남성"

# 두 번째 시트부터는 생성하기
sheet2_solo_women = wb.create_sheet(title="미혼여성")
sheet2_solo_women.column_dimensions["D"].width = 70

sheet3_married_women = wb.create_sheet(title="기혼여성")
sheet3_married_women.column_dimensions["D"].width = 70

sheet4_others = wb.create_sheet(title="기타")
sheet4_others.column_dimensions["D"].width = 70

# 생존자수/사망자 수 시트 생성
sheet5_report = wb.create_sheet(title="보고서")
sheet5_report.append(["분류", "생존자수", "사망자수", "생존율"])

# 성별 사망자수와 생존자 수 변수 선언
man_survived, man_unsurvived = 0, 0
single_survived, single_unsurvived = 0, 0
married_survived, married_unsurvived = 0, 0
others_survived, others_unsurvived = 0, 0


# 정규식
pattern = re.compile(r" [A-z]+\.")

for each_row in sheet1.rows:
    # 정규식과 매치된 부분 찾기
    data = pattern.findall(each_row[3].value)

    # train.xlsx의 첫번째 행에 있는 항목 이름 복사후 각 시트로 붙여넣기
    if each_row[0].row == 1:
        sheet1_man.append([each_item.value for each_item in each_row])
        sheet2_solo_women.append([each_item.value for each_item in each_row])
        sheet3_married_women.append([each_item.value for each_item in each_row])
        sheet4_others.append([each_item.value for each_item in each_row])
    else:  # 정규식 매치된 부분 각 시트에 옮기기
        if len(data) > 0:
            if data[0] == " Mr.":
                sheet1_man.append([each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    man_survived += 1
                else:
                    man_unsurvived += 1

            elif data[0] == " Miss.":
                sheet2_solo_women.append(
                    [each_item.value for each_item in each_row]
                )
                if each_row[1].value == 1:
                    single_survived += 1
                else:
                    single_unsurvived += 1
            elif data[0] == " Mrs.":
                sheet3_married_women.append(
                    [each_item.value for each_item in each_row]
                )
                if each_row[1].value == 1:
                    married_survived += 1
                else:
                    married_unsurvived += 1
            else:
                sheet4_others.append(
                    [each_item.value for each_item in each_row]
                )
                if each_row[1].value == 1:
                    others_survived += 1
                else:
                    others_unsurvived += 1

# 생존율 구하기
# 남성
man_survived_rate = "%.2f%%" % (
    man_survived / (man_survived + man_unsurvived) * 100
)
sheet5_report.append(["남성", man_survived, man_unsurvived, man_survived_rate])
# 미혼여성
single_survived_rate = "%.2f%%" % (
    single_survived / (single_survived + single_unsurvived) * 100
)
sheet5_report.append(
    ["미혼여성", single_survived, single_unsurvived, single_survived_rate]
)
# 기혼여성
married_survived_rate = "%.2f%%" % (
    married_survived / (married_survived + married_unsurvived) * 100
)
sheet5_report.append(
    ["기혼여성", married_survived, married_unsurvived, married_survived_rate]
)
# 기타
others_survived_rate = "%.2f%%" % (
    others_survived / (others_survived + others_unsurvived) * 100
)
sheet5_report.append(
    ["기타", others_survived, others_unsurvived, others_survived_rate]
)


wb.save("./resources/train_gender.xlsx")
wb.close()
excel_file.close()
